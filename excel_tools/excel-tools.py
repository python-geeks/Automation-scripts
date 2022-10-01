# python script that deals with basic excel operations
# it is better to keep the excel file in the same folder
# enter full address carefully if you are doing it manually


# download and install openpyxl
import openpyxl
print("enter the name of the excel file")
path = input()
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


print("if you want to get value from cell type OUT")
if input() == "OUT":
    print("remember that cells start from 1,1")
    print("enter row, column eg 5,1")
    m, n = list(map(int, input().split()))
    cell_obj = sheet_obj.cell(row=m, column=n)
    print("the value of the given cell is: ", cell_obj.value)


print("if you want the entire excel file displayed type OUTX")
if input() == "OUTX":
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    for i in range(1, row + 1):
        for j in range(1, column + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            print(cell_obj.value)
print("if you want to create a new excel file, print NEW")
if input() == "NEW":
    print("enter your workbook name eg name.xlsx")
    s = input()
    from openpyxl import Workbook
    wbk = Workbook()
    wbk.save(filename=s)
    sheet = wbk.active
    e = ""
    while e != "NO":
        print("enter row, column of cell in which you want to input data")
        a, b = list(map(int, input().split()))
        print("enter value")
        t = input()
        v = sheet.cell(row=m, column=n)
        v.value = t
        wbk.save(s)
        print("type NO to exit, YES to repeat")
        e = input()
        print("type MERGE to merge desired cells")
if input() == "MERGE":
    print("enter range to be merged like C1:F6")
    p = input()
    sheet = wbk.active
    sheet.merge_cells(p)
    wbk.save(s)
print("to copy from one file to another type COPY")
if input() == "COPY":
    print("enter source file address")
    filename = input()
    wb1 = openpyxl.load_workbook(filename)
    ws1 = wb1.worksheets[0]
    print("enter destination file address")
    filename1 = input()
    wb2 = openpyxl.load_workbook(filename1)
    ws2 = wb2.active
    for i in range(1, row + 1):
        for j in range(1, column + 1):
            c = ws1.cell(row=i, column=j)
            wb2.cell(row=i, column=j).value = c.value
    wb2.save(str(filename1))
print("if you want to create a bar chart type CHART")
if input() == "CHART":
    from openpyxl.chart import BarChart3D, Reference
    print("how many columns do you want")
    lel = int(input())
    print("enter", lel, " values")
    for i in range(lel):
        b = int(input())
        sheet.append([b])
    values = Reference(sheet, min_col=1, min_row=1,
                       max_col=1, max_row=lel)
    chart = BarChart3D()
    chart.add_data(values)
    print("enter chart title")
    title = input()
    chart.title = title
    chart.x_axis.title = " X AXIS "
    chart.y_axis.title = " Y AXIS "
    sheet.add_chart(chart, "E2")
    wbk.save("BarChart.xlsx")
