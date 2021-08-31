import socket
import tqdm
import os
import extract
import excel
import datetime
import shutil
import dynamic_fac
import time
def compute():
    # device's IP address
    SERVER_HOST = "0.0.0.0"
    SERVER_PORT = 5001
    # receive 4096 bytes each time
    BUFFER_SIZE = 4096
    SEPARATOR = "<SEPARATOR>"

    s = socket.socket()
    try:
        s.bind((SERVER_HOST, SERVER_PORT))
    except:
        o="Unable to bind"
        return o
    while True:
        s.listen(5)
        project=os.getcwd()+"/"             # to get the current path of server file
        default=project+"images/"           # Adding static image file path to above project folder path
        if not os.path.isdir(str(default)): # checking for static/images folder if not exists creating it and its inner folders of error,extracted, excelor _ in progress:
            os.mkdir(str(default))
        erpath=str(default+"error/")
        if not os.path.isdir(erpath):
            os.mkdir(erpath)
        excelpath=str(default+"excel/")
        if not os.path.isdir(excelpath):
            os.mkdir(excelpath)
        opath=str(default+"original/")
        if not os.path.isdir(opath):
            os.mkdir(str(default+"original"))
        dailypath=str(default+"daily/")
        if not os.path.isdir(dailypath):
            os.mkdir(dailypath)
        temppath=str(default+"temporary/")
        if not os.path.isdir(temppath):
            os.mkdir(temppath)
        jsonpath=str(default+"/original.json")

        # Creating the image file name
        foldername=str(datetime.date.today())
        filename=""
        print(filename)
        i=1
        while(i):
            todayfolder=str(dailypath+foldername+"/")
            print(todayfolder)
            if not os.path.isdir(todayfolder):
                os.mkdir(todayfolder)
            fpath1=todayfolder+"/"+str(i)+".jpg"
            if(os.path.isfile(fpath1)):
                i+=1
                print("in if"+filename)
                continue
            else:
                filename=fpath1
                print("in else"+filename)
                tempfile = str(i)
                print("tgis is temp file"+tempfile)
                break
        print(filename)
        print(f"[*] Listening as {SERVER_HOST}:{SERVER_PORT}")
        # accept connection if there is any
        client_socket, address = s.accept() 
        # if below code is executed, that means the sender is connected
        print(f"[+] {address} is connected.")
        # receive the file infos
        # receive using client socket, not server socket
        recd = client_socket.recv(BUFFER_SIZE)
        try:
            received = recd.decode()
        except:
            u="Unable to recieve, Please retake image"
            print(u)
            return u
        fname, filesize = received.split(SEPARATOR)
        # convert to integer
        filesize = int(filesize)
        # start receiving the file from the socket
        # and writing to the file stream
        progress = tqdm.tqdm(range(filesize), f"Receiving {filename}", unit="B", unit_scale=True, unit_divisor=1024)
        with open(filename, "wb") as f:
            for _ in progress:
                # read 1024 bytes fromsg.encode()m the socket (receive)
                bytes_read = client_socket.recv(BUFFER_SIZE)
                if not bytes_read:    
                    # nothing is received
                    # file transmitting is done
                    break
                f.write(bytes_read)
                progress.update(len(bytes_read))
        print(filename)
        extract.extract_count(filename,temppath,tempfile)
        print(filename)
        #shutil.copy(str(filename),str(temppath))
        l=dynamic_fac.match(filename,temppath,erpath,jsonpath,excelpath)
        print(filename)
        msg = excel.create_sheet(l,excelpath)
        print("number of faces matched : ",len(l))
        print("Attendance marked for: ",l)
        break
        # close the client socket
        #client_socket.close()
        # close the server socket
    s.close()
    print("server closed")
    return(msg)
def sendmail():
    import smtplib 
    from email.mime.multipart import MIMEMultipart 
    from email.mime.text import MIMEText 
    from email.mime.base import MIMEBase 
    from email import encoders
    from datetime import datetime,date
    import openpyxl
    import pymongo
    client=pymongo.MongoClient('mongodb://127.0.0.1:27017/FacultyLogin')
    mydb=client['FacultyLogin']
    user =mydb['users']
    send_mail=user.find_one({"username":"ADMIN" })
    #print(send_mail)
    #print(send_mail["mailid"])
    user_mail = send_mail["mailid"]
    #print(user_mail)
    while True:
        now = datetime.now()
        current_time=now.strftime("%H:%M:%S")
        if current_time=="11:00:00" or current_time=="18:00:00":
            fromaddr = "attendance@nmrec.edu.in"
            toaddr = user_mail

            # instance of MIMEMultipart 
            msg = MIMEMultipart() 

            # storing the senders email address 
            msg['From'] = fromaddr 

            # storing the receivers email address 
            msg['To'] = toaddr 

            # storing the subject 
            msg['Subject'] = "Subject of the Mail"

            # string to store the body of the mail 
            body = "Body_of_the_mail"

            # attach the body with the msg instance 
            msg.attach(MIMEText(body, 'plain')) 
            day=int(date.today().day)
            mnth=str(date.today().month)
            if(int(mnth)<10):
                mnth=str('0'+mnth)
            yr=str(date.today().year)
            today = str(date.today())
            # open the file to be sent 
            try:
                print("intry")
                if day<=15:
                    filename = os.getcwd()+"/images/excel/"+yr+"-"+mnth+"-"+"01.xlsx"
                else:
                    filename = os.getcwd()+"/images/excel/"+yr+"-"+mnth+"-"+"16.xlsx"
                wb=openpyxl.load_workbook(filename)
                ws=wb.active
                ws=wb[today]
                mrow=ws.max_row
                mcol=ws.max_column
                wb_temp=openpyxl.Workbook()
                ws_temp=wb_temp.active
                for i in mrow:
                    for j in mcol:
                        ws_temp.cell(row=i,column=j).value=ws.cell(row=i,column=j).value
                wb_temp.save(os.getcwd()+"/images/excel/tempfile.xlsx")
                variable=os.getcwd()+"/images/excel/tempfile.xlsx"
                attachment = open(variable, "rb")
            except:
                print("in except")
                wb=openpyxl.Workbook()
                ws=wb.active
                wb.save(os.getcwd()+"/images/excel/Empty_file.xlsx")
                empty=os.getcwd()+"/images/excel/Empty_file.xlsx"
                attachment = open(empty,"rb")
                
            # instance of MIMEBase and named as p 
            p = MIMEBase('application', 'octet-stream') 

            # To change the payload into encoded form 
            p.set_payload((attachment).read()) 

            # encode into base64 
            encoders.encode_base64(p) 

            p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 

            # attach the instance 'p' to instance 'msg' 
            msg.attach(p) 

            # creates SMTP session 
            s = smtplib.SMTP('smtp.gmail.com', 587) 

            # start TLS for security 
            s.starttls() 

            # Authentication 
            s.login(fromaddr, "nmrec@frba") 

            # Converts the Multipart msg into a string 
            text = msg.as_string() 

            # sending the mail 
            s.sendmail(fromaddr, toaddr, text) 
            
            # terminating the session 
            s.quit()
            print ("Email sent at:")
            print(current_time) 
            break
        else:
            break
while True:
    msg = compute()
    message = str.encode(msg)
    import reciever
    reciever.re(message,x)
    sendmail()