import openpyxl
from openpyxl.styles import Font 
from datetime import date,datetime
import gmail
day=int(date.today().day)
mnth=str(date.today().month)
if(int(mnth)<10):
    mnth=str('0'+mnth)
yr=str(date.today().year)
today = str(date.today())

def compute(ws,day,c,l,wb,excelpath):
    name = []
    email = []
    print("called compute")
    now = datetime.now()
    time=str(now.strftime("%I:%M %p"))
    ws=wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.row_dimensions[1].height = 38
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    max_col = ws.max_column 
    m_row = ws.max_row
    if(ws.cell(row = 1,column = max_col+1).value == None):
        ws.cell(row =1 ,column = max_col+1).value = "IN-TIME"    
        ws.cell(row =1 ,column = max_col+2).value = "OUT-TIME"
    for i in range(2, m_row + 1):
        for j in l:
            if(ws.cell(row = i, column = 1).value == j):
                while True:
                    if(ws.cell(row = i , column = c+1).value == None):
                        ws.cell(row = i, column = c+1).value = time
                        na = ws.cell(row = i, column = 2).value
                        name.append(na)
                        Mail = str(ws.cell(row = i, column = 3).value)
                        email.append(Mail)
                        break
                    else:
                        c+=1
    if(day > 0 and day < 16):
        wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        #create_excelhtml.view_attendance(cur_sheetpath)
        #l=create_excelhtml.view_absentees(cur_sheetpath,c) # c=h+1
    else:
        wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        #create_excelhtml.view_attendance(cur_sheetpath)
        #l=create_excelhtml.view_absentees(cur_sheetpath,c) # c=h+1
    if l:
        i=0
        for m in email:
            gmail.mail(m,name[i])
            i+=1
        msg = str(','.join(l))+':'+str(','.join(name))
    else:
        msg = "Image not found!,Please retake image"
    print(msg)
    return msg
def create_sheet(l,excelpath):    
    #creating a new excel sheet
    if(day == 1 or day == 16):
        print("new excel sheet")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = today
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        m_row = ws_temp.max_row
        for i in range(1, m_row + 1):
            for j in range(1,m_column+1):
                ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
        wb.save(excelpath+today+".xlsx")
        l=compute(ws,day,m_column,l,wb,excelpath)
     
    #loading an excelsheet if date < 16 
     
    elif(day > 1 and day < 16):
        print('date < 16')
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        l=compute(ws,day,m_column,l,wb,excelpath)
        
    #loading a excel sheet if date  > 16    
        
    else:
        print("date > 16")
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames 
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        l=compute(ws,day,m_column,l,wb,excelpath)
    return l






