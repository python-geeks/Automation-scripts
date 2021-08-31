import socket
import tqdm
import os
import time
import multiprocessing
import openpyxl
import threading
from openpyxl import Workbook
import openpyxl as xl; 
import schedule
def recvr(h,p):
    SERVER_HOST = h
    SERVER_PORT = p
    if p==5001:
        pi=1
    if p==5003:
        pi=2
    BUFFER_SIZE = 4096
    SEPARATOR = "<SEPARATOR>"

    s = socket.socket()
    try:
        s.bind((SERVER_HOST, SERVER_PORT))
    except:
        o="Unable to bind"
        return o
    while True:
        s.listen(5)
        print(f"[*] Listening as {SERVER_HOST}:{SERVER_PORT}")
        client_socket, address = s.accept() 
        print(f"[+] {address} is connected.")
        recd = client_socket.recv(BUFFER_SIZE)
        try:
            received = recd.decode()
        except:
            u="Unable to recieve, Please retake image"
            print(u)
            return u
        filename, filesize = received.split(SEPARATOR)
        filesize = int(filesize)
        progress = tqdm.tqdm(range(filesize), f"Receiving {filename}", unit="B", unit_scale=True, unit_divisor=1024)
        with open(filename, "wb") as f:
            for _ in progress:
                bytes_read = client_socket.recv(BUFFER_SIZE)
                if not bytes_read:    
                    break
                f.write(bytes_read)
                progress.update(len(bytes_read))
        break
    s.close()
    print("server closed")
    time.sleep(1)
    #sendr()
    # if pi==1:
    #     time.sleep(0.1)
    #     s1.start()
    #     time.sleep(0.1)
    # if pi==2:
    #     time.sleep(0.1)
    #     s2.start()
        
    #     time.sleep(0.1)
def sendr():
    SEPARATOR = "<SEPARATOR>"
    BUFFER_SIZE = 4096 # send 4096 bytes each time step
    # the ip address or hostname of the server, the receiver
    host = "192.168.0.104"
    # the port, let's use 5001
    port = 5003
    # the name of file we want to send, make sure it exists
    filename = "source.xlsx"
    # get the file size
    filesize = os.path.getsize(filename)
    # create the client socket
    s= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    print(f"[+] Connecting to {host}:{port}")
    s.connect((host, port))
    print("[+] Connected.")
    # send the filename and filesize
    s.send(f"{filename}{SEPARATOR}{filesize}".encode())
    # start sending the file
    progress = tqdm.tqdm(range(filesize), f"Sending {filename}", unit="B", unit_scale=True, unit_divisor=1024)
    with open(filename, "rb") as f:
        for _ in progress:
            # read the bytes from the file
            bytes_read = f.read(BUFFER_SIZE)
            if not bytes_read:
                # file transmitting is done
                break
            # we use sendall to assure transimission in 
            # busy networks
            s.sendall(bytes_read)
            # update the progress bar
            progress.update(len(bytes_read))
    s.close()
def sendr2():
    SEPARATOR = "<SEPARATOR>"
    BUFFER_SIZE = 4096 # send 4096 bytes each time step
    # the ip address or hostname of the server, the receiver
    host = "192.168.0.104"
    # the port, let's use 5001
    port = 5005
    # the name of file we want to send, make sure it exists
    filename = "source.xlsx"
    # get the file size
    filesize = os.path.getsize(filename)
    # create the client socket
    s= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    print(f"[+] Connecting to {host}:{port}")
    s.connect((host, port))
    print("[+] Connected.")
    # send the filename and filesize
    s.send(f"{filename}{SEPARATOR}{filesize}".encode())
    # start sending the file
    progress = tqdm.tqdm(range(filesize), f"Sending {filename}", unit="B", unit_scale=True, unit_divisor=1024)
    with open(filename, "rb") as f:
        for _ in progress:
            # read the bytes from the file
            bytes_read = f.read(BUFFER_SIZE)
            if not bytes_read:
                # file transmitting is done
                break
            # we use sendall to assure transimission in 
            # busy networks
            s.sendall(bytes_read)
            # update the progress bar
            progress.update(len(bytes_read))
    s.close()
def serr():
    workbook = Workbook()
    sheet = workbook.active
    workbook.save(filename="source.xlsx")
    filename ="template.xlsx"
    wb1 = xl.load_workbook(filename) 
    ws1 = wb1.worksheets[0] 
    filename1 ="source.xlsx"
    wb2 = xl.load_workbook(filename1) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column  
    for i in range (1, mr + 1): 
        for j in range (1, mc + 1): 
            c = ws1.cell(row = i, column = j)  
            ws2.cell(row = i, column = j).value = c.value 
    wb2.save(str(filename1)) 
    wb0 = openpyxl.load_workbook('Emp.xlsx')
    ws0 = wb0.active
    wb1 = openpyxl.load_workbook('Emp2.xlsx')
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook('source.xlsx')
    ws2 = wb2.active
    for src, dst in zip(ws1['D:D'], ws2['D:D']):
        dst.value = src.value
    for src, dst in zip(ws0['D:D'], ws2['D:D']):
        if not dst.value:
            dst.value = src.value
    for src, dst in zip(ws1['E:E'], ws2['E:E']):
        dst.value = src.value
    for src, dst in zip(ws0['E:E'], ws2['E:E']):
        if not dst.value:
            dst.value = src.value
    wb2.save('source.xlsx') 
    time.sleep(0.5)
    #os.remove("source.xlsx")
h1='0.0.0.0'
h2='0.0.0.0'
p1=5001
p2=5002
su1=5003
su2=5005
p1 = multiprocessing.Process(target=recvr, args=(h1,p1,)) 
p2 = multiprocessing.Process(target=recvr, args=(h2,p2,))
s1 = multiprocessing.Process(target=sendr, args=(su1,)) 
s2 = multiprocessing.Process(target=sendr, args=(su2,))

t1 = threading.Thread(target=sendr)
t2 = threading.Thread(target=sendr2)
def main():
    p1.start() 
    time.sleep(0.1)
    p2.start()
    time.sleep(0.1) 
    time.sleep(0.1)
    p1.join()
    p2.join()
    serr()
    t1.start()
    time.sleep(0.5)
    t2.start()
    time.sleep(0.5)
    t1.join()
    time.sleep(0.5)
    t2.join()
    time.sleep(0.5)
if __name__ == "__main__":
    multiprocessing.freeze_support()
    schedule.every().day.at("10:23").do(main)
    schedule.every().day.at("17:52").do(main)
    while True:
        schedule.run_pending()