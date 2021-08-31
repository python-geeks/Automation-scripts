import RPi.GPIO as GPIO
import time
import picamera
import face_recognition.api as face_recognition
import os,shutil
import json
import numpy
from PIL import Image
from datetime import date,datetime
import smtplib
import random
import math
import openpyxl
from openpyxl.styles import Font 
import openpyxl
from openpyxl.styles import Font 
import cv2 
import shutil
from time import sleep 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders
import tkinter as tk
from PIL import Image,ImageEnhance
import face_recognition
from resizeimage import resizeimage
import tflite_runtime.interpreter as tflite
import argparse
import numpy as np
import sys
from threading import Thread
import importlib.util
import schedule 
import threading
import socket
import tqdm
import shutil
day=int(date.today().day) 
def extract_count(gpath,epath,filename):
    print("in extract")
    image = face_recognition.load_image_file(gpath)
    face_locations = face_recognition.face_locations(image,number_of_times_to_upsample=0)
    count=len(face_locations)
    if count==0:
        img=Image.open(gpath)
        l=img.size
        m=max(l)
        img.save(gpath,img.format)
        out=Image.open(gpath)
        out=out.rotate(270)
        out.save(gpath)
        image = face_recognition.load_image_file(gpath)
        face_locations = face_recognition.face_locations(image)
        count=len(face_locations)
    if count>0:
        i=0
        for face_location in face_locations:
            i=i+1
            top, right, bottom, left = face_location
            face_image = image[top:bottom, left:right]
            pil_image = Image.fromarray(face_image)
            name =filename+ "-"+str(i) +'.png'
            pil_image.save(epath+name)
    print("exiting extract")
    return count 
def match(current_img,temppath,erpath,jsonpath,excelpath):
    print("in match")
    now = datetime.now()
    time=str(now.strftime("%I:%M %p"))
    data = open(jsonpath,"r")
    n=data.read()
    originallist=json.loads(n)
    for i in originallist:
        originallist[i]=numpy.asarray(originallist[i])
    l=[]
    
    for a in range(0,2):
        extractedlist=[]
        for filename in os.listdir(temppath):
            sublist=[]
            filename=str(filename)
            sublist.append(filename)
            file_path=temppath+filename
            image=face_recognition.load_image_file(file_path)
            image_encoding=face_recognition.face_encodings(image,num_jitters=5)
            if len(image_encoding)>0:
                image_encoding=image_encoding[0]
                sublist.append(image_encoding)
                extractedlist.append(sublist)
            else:
                shutil.move(temppath+filename,erpath)
                se1("Error! Please retake the photo")
        for key in originallist:
            rollno=key[0:20]
            if rollno in l:
                continue
            for esublist in extractedlist:
                oimg=originallist[key]
                eimg=esublist[1]
                result=face_recognition.compare_faces([oimg],eimg,tolerance=0.435)
                if result[0]== True:
                    os.remove(temppath+esublist[0])
                    l.append(rollno)
                    extractedlist.remove(esublist)
                    break
    for temp_img in os.listdir(temppath):
        os.remove(temppath+str(temp_img))
    for err_img in os.listdir(erpath):
        os.remove(erpath+str(err_img))
    print("exiting match")
    return l
def mail(email,name):
    print("in mail")
    now = datetime.now()
    time=str(now.strftime("%I:%M %p"))
    content = '\nHello '+name+'\n Thank you using Face recognition based attendance,\n Your timestamp has been recorded at '+time+'.\n Thank you' 
    username = "attendance@nmrec.edu.in"
    password = "nmrec@frba"
    sender = "attendance@nmrec.edu.in"
    recipient = email
    mail = smtplib.SMTP("smtp.gmail.com",587)
    mail.ehlo() 
    mail.starttls() 
    mail.ehlo()
    mail.login(username,password)
    header = 'To:' + recipient + '\n' + 'From:' + sender + '\n' + 'Subject: Time Stamp Recorded \n'
    content = header+content
    mail.sendmail(sender,recipient,content)
    mail.close
    print("exiting mail")
    return(0)
def comput(ws,day,c,l,wb,excelpath):
    print("in comput")
    day=int(date.today().day)
    mnth=str(date.today().month)
    if(int(mnth)<10):
        mnth=str('0'+mnth)
    yr=str(date.today().year)
    today = str(date.today())
    name = []
    email = []
    now = datetime.now()
    time=str(now.strftime("%I:%M %p"))
    ws=wb.active
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.row_dimensions[1].height = 38
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    max_col = ws.max_column 
    m_row = ws.max_row
    if(ws.cell(row = 1,column = max_col+1).value == None):
        ws.cell(row =1 ,column = max_col+1).value = "IN-TIME"    
        ws.cell(row =1 ,column = max_col+2).value = "OUT-TIME"
    for i in range(2, m_row + 1):
        for j in l:
            if(ws.cell(row = i, column = 1).value == j):
                while True:
                    if(ws.cell(row = i , column = c+1).value == None):
                        ws.cell(row = i, column = c+1).value = time
                        na = ws.cell(row = i, column = 2).value
                        name.append(na)
                        Mail = str(ws.cell(row = i, column = 3).value)
                        email.append(Mail)
                        break
                    else:
                        c+=1
    if(day > 0 and day < 16):
        wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
    else:
        wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
        cur_sheetpath=str(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
    if l:
        i=0
        for m in email:
            t=threading.Thread(target=mail, args=(m,name[i]))
            t.start()
            t.join()
            i+=1
        msg = str(','.join(l))+':'+str(','.join(name))
    else:
        msg = "Image not found!,Please retake image"
    print(msg)
    print("exiting comput")
    return msg
def create_sheet(l,excelpath):
    print("in create sheet")
    day=int(date.today().day)
    mnth=str(date.today().month)
    if(int(mnth)<10):
        mnth=str('0'+mnth)
    yr=str(date.today().year)
    today = str(date.today())    
    if(day == 1 or day == 16):
        print("new excel sheet")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = today
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        m_row = ws_temp.max_row
        for i in range(1, m_row + 1):
            for j in range(1,m_column+1):
                ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
        wb.save(excelpath+today+".xlsx")
        l=comput(ws,day,m_column,l,wb,excelpath)
      
     
    elif(day > 1 and day < 16):
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value
            wb.save(excelpath+yr+"-"+mnth+"-"+"01"+".xlsx")
        l=comput(ws,day,m_column,l,wb,excelpath)    
        
    else:
        try:
            wb = openpyxl.load_workbook(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx")
            ws = wb.active
        except:
            wb = openpyxl.Workbook()
            ws = wb.active
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        wb_temp = openpyxl.load_workbook(excelpath+"/template.xlsx")
        ws_temp = wb_temp.active
        m_column = ws_temp.max_column
        list = wb.sheetnames 
        if(today in list):
            ws = wb[today]
        else:
            wb.create_sheet(index = 0 , title = today)
            ws = wb.active
            m_row = ws_temp.max_row 
            for i in range(1, m_row + 1):
                for j in range(1,m_column+1):
                    ws.cell(row = i, column = j).value = ws_temp.cell(row = i, column = j).value 
            wb.save(excelpath+yr+"-"+mnth+"-"+"16"+".xlsx") 
        l=comput(ws,day,m_column,l,wb,excelpath)
    print("exiting createsheet")
    return l
def compute():
    face_cascade = cv2.CascadeClassifier('Cascades/haarcascade_frontalface_default.xml')
    project=os.getcwd()+"/"             
    default=project+"images/"           
    if not os.path.isdir(str(default)): 
        os.mkdir(str(default))
    erpath=str(default+"error/")
    if not os.path.isdir(erpath):
        os.mkdir(erpath)
    excelpath=str(default+"excel/")
    if not os.path.isdir(excelpath):
        os.mkdir(excelpath)
    opath=str(default+"original/")
    if not os.path.isdir(opath):
        os.mkdir(str(default+"original/"))
    dailypath=str(default+"daily/")
    if not os.path.isdir(dailypath):
        os.mkdir(dailypath)
    temppath=str(default+"temporary/")
    if not os.path.isdir(temppath):
        os.mkdir(temppath)
    jsonpath=str(default+"/original.json")
    foldername=str(date.today())
    filename=""
    i=1
    while(i):
        todayfolder=str(dailypath+foldername+"/")
        if not os.path.isdir(todayfolder):
            os.mkdir(todayfolder)
        fpath1=todayfolder+"/"+str(i)+".png"
        if(os.path.isfile(fpath1)):
            i+=1
            continue
        else:
            filename=fpath1
            tempfile = str(i)
            break
    camera=PiCamera()
    camera.start_preview()
    sleep(3)        
    camera.capture('test.png')
    camera.stop_preview()
    camera.close()
    im=Image.open('test.png')
    width, height=im.size
    left= (width-600)/2
    top=(height-600)/2
    right=(width+600)/2
    bottom=(height+720)/2
    im=im.crop((left,top,right,bottom))
    im.save('/'+todayfolder+'/'+str(i)+'.png')
    hi=tf()
    if hi==1:
        f="please retake image"
        return(f)
    else:
        extract_count(filename,temppath,tempfile)
        l=match(filename,temppath,erpath,jsonpath,excelpath)
        msg = create_sheet(l,excelpath)
        se(msg)
def se(ms):
    root=tk.Tk()
    t=str(ms)
    t.replace(':',':\n')
    x=t.split(':')            
    if x[0]:
        try:
            
            l=tk.Label(root,text="Time stamp recorded for:\n"+x[0]+'\n'+x[1],font=("Times New Roman",28),fg='red')
            l.place(relx=0.5,rely=0.5,anchor='center')
        except:
            l=tk.Label(root,text=t,font=("Times New Roman",28),fg='red')
            l.place(relx=0.5,rely=0.5,anchor='center')
    else:
        l=tk.Label(root,text=t,font=("Times New Roman",28),fg='red')
        l.place(relx=0.5,rely=0.5,anchor='center')
    root.attributes('-fullscreen',True)
    if x[0]:
        root.after(5000,lambda: root.destroy())
    root.mainloop()
    return(0)
def se1(ms):
    root=tk.Tk()
    t=str(ms)          
    if t.find("ensor"):
        try:
            x=StringVar()
            x.set(x.get()+'\n Please move towards Sensor')
            l=tk.Label(root,textvariable=x,font=("Times New Roman",28),fg='red')
            l.place(relx=0.5,rely=0.5,anchor='center')
        except:
            l=tk.Label(root,text=t,font=("Times New Roman",28),fg='red')
            l.place(relx=0.5,rely=0.5,anchor='center')
    else:
        l=tk.Label(root,text=t,font=("Times New Roman",28),fg='red')
        l.place(relx=0.5,rely=0.5,anchor='center')
    root.attributes('-fullscreen',True)
    root.after(1000,lambda: root.destroy())
    root.mainloop()
    return(0)
def sendmail():
    try:
        while True:
            fromaddr = "attendance@nmrec.edu.in"
            toaddr = "18b61a05d6@nmrec.edu.in"
            msg = MIMEMultipart() 
            msg['From'] = fromaddr 
            msg['To'] = toaddr 
            msg['Subject'] = "Subject of the Mail"

            body = "Body_of_the_mail"
            msg.attach(MIMEText(body, 'plain')) 
            day=int(date.today().day)
            mnth=str(date.today().month)
            if(int(mnth)<10):
                mnth=str('0'+mnth)
            yr=str(date.today().year)
            today = str(date.today()) 
            print("intry")
            print(yr+"-"+mnth+"-")
            if day<=15:
                filename = os.getcwd()+"/images/excel/"+yr+"-"+mnth+"-"+"01.xlsx"
            else:
                filename = os.getcwd()+"/images/excel/"+yr+"-"+mnth+"-"+"16.xlsx"
            import pandas as pd
            empty=os.getcwd()+"/images/excel/Empty_File.xlsx"
            data=pd.read_excel(filename,sheet_name=today)
            data.to_excel(empty,sheet_name='temporary')
            attachment = open(empty,"rb")
            p = MIMEBase('application', 'octet-stream')
            p.set_payload((attachment).read())
            encoders.encode_base64(p) 

            p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 

            msg.attach(p) 

            s = smtplib.SMTP('smtp.gmail.com', 587) 
            s.starttls() 
            s.login(fromaddr, "nmrec@frba") 
            text = msg.as_string() 
            s.sendmail(fromaddr, toaddr, text) 
            s.quit()
            print ("Email sent at:")
            print(current_time) 
            break
    except:
        print("There is no "+today+" worksheet")
def tf():
    print("In object")        
    parser = argparse.ArgumentParser()
    parser.add_argument('--model', help='Provide the path to the TFLite file, default is models/model.tflite',
                        default='models/model.tflite')
    parser.add_argument('--labels', help='Provide the path to the Labels, default is models/labels.txt',
                        default='models/labels.txt')
    parser.add_argument('--image', help='Name of the single image to perform detection on',
                        default='test.png')
    parser.add_argument('--threshold', help='Minimum confidence threshold for displaying detected objects',
                        default=0.5)
                        
    args = parser.parse_args()                              

    PATH_TO_MODEL_DIR = args.model
    PATH_TO_LABELS = args.labels
    IMAGE_PATH = args.image
    MIN_CONF_THRESH = float(args.threshold)    
    start_time = time.time()
    interpreter = tflite.Interpreter(model_path=PATH_TO_MODEL_DIR)
    with open(PATH_TO_LABELS, 'r') as f:
        labels = [line.strip() for line in f.readlines()]
    end_time = time.time()
    elapsed_time = end_time - start_time

    interpreter.allocate_tensors()
    input_details = interpreter.get_input_details()
    output_details = interpreter.get_output_details()

    height = input_details[0]['shape'][1]
    width = input_details[0]['shape'][2]

    floating_model = (input_details[0]['dtype'] == np.float32)

    input_mean = 127.5
    input_std = 127.5
    frame_rate_calc = 1
    freq = cv2.getTickFrequency()
    current_count=0
    t1 = cv2.getTickCount()
    image = cv2.imread(IMAGE_PATH)
    image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    imH, imW, _ = image.shape
    image_resized = cv2.resize(image_rgb, (width, height))
    input_data = np.expand_dims(image_resized, axis=0)

    if floating_model:
        input_data = (np.float32(input_data) - input_mean) / input_std

    interpreter.set_tensor(input_details[0]['index'],input_data)
    interpreter.invoke()
    boxes = interpreter.get_tensor(output_details[0]['index'])[0]
    classes = interpreter.get_tensor(output_details[1]['index'])[0]
    scores = interpreter.get_tensor(output_details[2]['index'])[0]
    for i in range(len(scores)):
        if ((scores[i] > MIN_CONF_THRESH) and (scores[i] <= 1.0)):
            ymin = int(max(1,(boxes[i][0] * imH)))
            xmin = int(max(1,(boxes[i][1] * imW)))
            ymax = int(min(imH,(boxes[i][2] * imH)))
            xmax = int(min(imW,(boxes[i][3] * imW)))
            
            cv2.rectangle(image, (xmin,ymin), (xmax,ymax), (10, 255, 0), 2)

            
            object_name = labels[int(classes[i])] 
            label = '%s: %d%%' % (object_name, int(scores[i]*100)) 
            current_count+=1
            print("Exiting objecdt")
            if(object_name=='phone' and object_name=='person') or (object_name=='paper' and object_name=='person') or (object_name=='ipad' and object_name=='person') or (object_name=='phone') or (object_name=='paper') or (object_name=='ipad'):
                return(1)
            else:
                return(0)
def sendr():
    if(day > 0 and day < 16):
        src=os.getcwd()+"/images/excel/2021-01-01.xlsx"
        dst=os.getcwd()
        shutil.copy(src, dst)
    else:
        src=os.getcwd()+"/images/excel/2021-01-16.xlsx"
        dst=os.getcwd() 
        shutil.copy(src, dst)
    SEPARATOR = "<SEPARATOR>"
    BUFFER_SIZE = 4096 
    host = "192.168.0.104" #Static ip of server(localserver)
    port = 5002 #device1 will have this port
    import openpyxl as xl; 
    if(day > 0 and day < 16):
        filename =os.getcwd()+"/2021-01-01.xlsx"
    else:
        filename =os.getcwd()+"/2021-01-16.xlsx"
    wb1 = xl.load_workbook(filename) 
    ws1 = wb1.worksheets[0] 
    filename1 =os.getcwd()+"/Emp2.xlsx"
    wb2 = xl.load_workbook(filename1) 
    ws2 = wb2.active 
    mr = ws1.max_row 
    mc = ws1.max_column 
    for i in range (1, mr + 1): 
        for j in range (1, mc + 1): 
            c = ws1.cell(row = i, column = j) 
            ws2.cell(row = i, column = j).value = c.value 
    wb2.save(str(filename1)) 
    filename = "Emp2.xlsx"
    filesize = os.path.getsize(filename)
    s= socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    print(f"[+] Connecting to {host}:{port}")
    s.connect((host, port))
    print("[+] Connected.")
    s.send(f"{filename}{SEPARATOR}{filesize}".encode())
    progress = tqdm.tqdm(range(filesize), f"Sending {filename}", unit="B", unit_scale=True, unit_divisor=1024)
    with open(filename, "rb") as f:
        for _ in progress:
            bytes_read = f.read(BUFFER_SIZE)
            if not bytes_read:
                break
            s.sendall(bytes_read)
            progress.update(len(bytes_read))
    s.close()
    recvr()
def recvr():
    SERVER_HOST = "0.0.0.0"
    SERVER_PORT = 5005
    BUFFER_SIZE = 4096
    SEPARATOR = "<SEPARATOR>"

    s = socket.socket()
    try:
        s.bind((SERVER_HOST, SERVER_PORT))
    except:
        o="Unable to bind"
        return o
    while True:
        s.listen(5)
        print(f"[*] Listening as {SERVER_HOST}:{SERVER_PORT}")
        client_socket, address = s.accept() 
        print(f"[+] {address} is connected.")
        recd = client_socket.recv(BUFFER_SIZE)
        try:
            received = recd.decode()
        except:
            u="Unable to recieve, Please retake image"
            print(u)
            return u
        filename, filesize = received.split(SEPARATOR)
        filesize = int(filesize)
        progress = tqdm.tqdm(range(filesize), f"Receiving {filename}", unit="B", unit_scale=True, unit_divisor=1024)
        with open(filename, "wb") as f:
            for _ in progress:
                bytes_read = client_socket.recv(BUFFER_SIZE)
                if not bytes_read:    
                    break
                f.write(bytes_read)
                progress.update(len(bytes_read))
        break
    s.close()
    print("server closed")
    if(day > 0 and day < 16):
        import openpyxl as xl; 
        filename ="source.xlsx"
        wb1 = xl.load_workbook(filename) 
        ws1 = wb1.worksheets[0] 
        if(day > 0 and day < 16):
            filename1 =os.getcwd()+"/images/excel/2021-01-01.xlsx"
        else:
            filename1 =os.getcwd()+"/images/excel/2021-01-16.xlsx"
        wb2 = xl.load_workbook(filename1) 
        ws2 = wb2.worksheets[0]  
        mr = ws1.max_row 
        mc = ws1.max_column 
        for i in range (1, mr + 1): 
            for j in range (1, mc + 1): 
                c = ws1.cell(row = i, column = j) 
                ws2.cell(row = i, column = j).value = c.value 
        wb2.save(str(filename1)) 
    else:
        import openpyxl as xl; 
        filename ="source.xlsx"
        wb1 = xl.load_workbook(filename) 
        ws1 = wb1.worksheets[0]  
        filename1 =os.getcwd()+"/images/excel/2021-01-16.xlsx"
        wb2 = xl.load_workbook(filename1) 
        ws2 = wb2.worksheets[0]  
        mr = ws1.max_row 
        mc = ws1.max_column  
        for i in range (1, mr + 1): 
            for j in range (1, mc + 1): 
                c = ws1.cell(row = i, column = j) 
                ws2.cell(row = i, column = j).value = c.value 
        wb2.save(str(filename1)) 

TRIG=21
ECHO=20
GPIO.setmode(GPIO.BCM)
def place():
    dailypath=os.getcwd()+"/images/daily/"
    shutil.rmtree(dailypath)
schedule.every().saturday.at("17:35").do(place)
schedule.every().day.at("11:00").do(sendr)
schedule.every().day.at("17:00").do(sendr)
schedule.every().day.at("11:02").do(sendmail)
schedule.every().day.at("17:02").do(sendmail)
now = datetime.now()
current_time=now.strftime("%H:%M:%S")
if (current_time=="16:55:00" or current_time=="11:00:00"):
    sendmail()
while True:
    se1("Distance measurement in progress")
    GPIO.setup(TRIG,GPIO.OUT)
    GPIO.setup(ECHO,GPIO.IN)
    GPIO.output(TRIG,False)
    time.sleep(0.2)
    GPIO.output(TRIG,True)
    time.sleep(0.00001)
    GPIO.output(TRIG,False)
    while GPIO.input(ECHO)==0:
        pulse_start=time.time()
    while GPIO.input(ECHO)==1:
        pulse_end=time.time()
    pulse_duration=pulse_end-pulse_start
    distance=pulse_duration*17150
    distance=round(distance,2)
    se1(distance)
    time.sleep(2) 
    if(distance <115):
        compute()
    else:
       schedule.run_pending()
       se1("\nPlease move towards the sensor")